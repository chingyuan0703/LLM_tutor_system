import ollama
import os
import json
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from collections import defaultdict
from langchain_community.document_loaders import PyPDFLoader
from langchain_ollama import OllamaEmbeddings
from langchain_chroma import Chroma
from langchain.docstore.document import Document
from langchain.text_splitter import RecursiveCharacterTextSplitter

# ------------------ 教材來源 ------------------

# 列出所有要讀取並建立向量資料庫的教材 PDF 檔名
PDF_LIST = [
    "Abraham Silberschatz Operating System Concepts.pdf",
    "Computer Organization and Design.pdf"
]

# 儲存向量資料庫的目錄與教材紀錄清單
DB_DIR = "./db"
DB_INFO_FILE = os.path.join(DB_DIR, "dbinfo.json")

# ------------------ 檢查是否有新教材 ------------------

# 取得上次已載入的教材檔名清單
def get_loaded_pdf_list():
    if os.path.exists(DB_INFO_FILE):
        with open(DB_INFO_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

# 儲存目前教材檔名清單
def save_loaded_pdf_list(pdf_list):
    os.makedirs(DB_DIR, exist_ok=True)
    with open(DB_INFO_FILE, "w", encoding="utf-8") as f:
        json.dump(pdf_list, f, ensure_ascii=False, indent=2)

# ------------------ 自訂章節切割函式 + 分段 ------------------

# 依章節名稱（Chapter X）分章，並使用文字切割器細分內容以適合模型處理
def split_by_chapter_and_chunk(documents, chunk_size=1000, chunk_overlap=100):
    splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=chunk_overlap)
    chapters = []
    current_chapter = []
    current_title = ""
    for doc in documents:
        text = doc.page_content
        lines = text.splitlines()
        for line in lines:
            if re.match(r'^\s*(Chapter|CHAPTER)\s+\d+', line):
                if current_chapter:
                    content = "\n".join(current_chapter)
                    chunks = splitter.split_text(content)
                    for chunk in chunks:
                        chapters.append(Document(page_content=chunk, metadata={"chapter": current_title}))
                    current_chapter = []
                current_title = line.strip()
            current_chapter.append(line.strip())
    if current_chapter:
        content = "\n".join(current_chapter)
        chunks = splitter.split_text(content)
        for chunk in chunks:
            chapters.append(Document(page_content=chunk, metadata={"chapter": current_title}))
    return chapters

# ------------------ 建立知識庫 ------------------

# 如果教材清單與上次不一致就需要重建資料庫
need_refresh = sorted(PDF_LIST) != sorted(get_loaded_pdf_list())

if need_refresh:
    print("🧠 偵測到有新教材變動，開始重新建立向量資料庫...")
    documents = []

    for pdf_file in PDF_LIST:
        if os.path.exists(pdf_file):
            print(f"📘 載入教材：{pdf_file}")
            loader = PyPDFLoader(pdf_file)
            loaded_docs = loader.load()
            for doc in loaded_docs:
                doc.metadata["source_file"] = pdf_file  # 標記每段資料的教材來源
            documents.extend(loaded_docs)
        else:
            print(f"⚠️ 找不到教材：{pdf_file}")

    # 使用章節分段與文字切割進行內容拆解
    docs = split_by_chapter_and_chunk(documents)
    print(f"✅ 已依章節分段為 {len(docs)} 筆")

    # 建立嵌入向量與向量資料庫
    embedding = OllamaEmbeddings(model="nomic-embed-text")
    vectordb = Chroma.from_documents(docs, embedding, persist_directory=DB_DIR)
    save_loaded_pdf_list(PDF_LIST)
else:
    print("✅ 教材未變動，直接載入資料庫")
    embedding = OllamaEmbeddings(model="nomic-embed-text")
    vectordb = Chroma(persist_directory=DB_DIR, embedding_function=embedding)

# 建立檢索器供後續問答查詢使用
retriever = vectordb.as_retriever()

# ------------------ 儲存對話紀錄 ------------------

# 每次使用者問答都儲存到 Excel（UTF-8）中，避免亂碼與資訊流失

def save_chat_xlsx(user_msg, bot_reply):
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"chat_log_{today_str}.xlsx"
    time_str = datetime.now().strftime("%H:%M:%S")

    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "使用者", "AI 回覆"])

        ws.append([time_str, user_msg, bot_reply])
        wb.save(filename)
    except PermissionError:
        print("⚠️ 檔案可能已開啟中，請關閉 Excel 後重試。")

# ------------------ LLM 對話處理 ------------------

# 建立初始提示與模型設定
messages = [{"role": "system", "content": "你是一位用繁體中文回答的家教老師，請用簡單方式講解問題。"}]
MODEL_NAME = 'llama3'
last_user = ""
last_reply = ""

# 對話主函式

def chat_with_ollama(prompt):
    global last_user, last_reply

    # 嘗試從提問中自動抓出使用者想指定哪本教材
    selected_pdf = None
    for pdf_file in PDF_LIST:
        base = os.path.splitext(os.path.basename(pdf_file))[0].lower()
        if base in prompt.lower():
            selected_pdf = pdf_file
            break

    # 執行檢索，若有教材指定則過濾對應段落
    if selected_pdf:
        print(f"🎯 只搜尋教材：{selected_pdf}")
        all_docs = retriever.invoke(prompt)
        related_docs = [doc for doc in all_docs if doc.metadata.get("source_file") == selected_pdf][:3]
        if not related_docs:
            related_docs = all_docs[:3]  # 若無結果則回退
    else:
        related_docs = retriever.invoke(prompt)[:3]

    # 整理出答案的上下文引用來源與段落
    context_chunks = []
    for doc in related_docs:
        source = doc.metadata.get("source_file", "未知教材")
        page = doc.metadata.get("page", "未知頁數")
        chunk = f"[來自：{source} 第 {page} 頁]\n{doc.page_content}"
        context_chunks.append(chunk)

    context_text = "\n---\n".join(context_chunks)

    # 包裝 prompt 給模型（含上下文）
    rag_prompt = f"""以下是教材內容摘要，請根據這些內容來回答問題：\n\n{context_text}\n\n使用者問題：{prompt}\n請用繁體中文詳細解釋，並舉例子說明。"""

    messages.append({"role": "user", "content": rag_prompt})

    try:
        response = ollama.chat(model=MODEL_NAME, messages=messages)
        reply = response['message']['content']
        messages.append({"role": "assistant", "content": reply})

        # 儲存對話與更新記憶
        save_chat_xlsx(prompt, reply)
        last_user = prompt
        last_reply = reply

        # 控制歷史訊息數量，避免過長
        if len(messages) > 20:
            messages.pop(1)

        return reply
    except Exception as e:
        return f"❌ 發生錯誤：{e}"

# ------------------ 主程式 ------------------

if __name__ == "__main__":
    print("📘 家教系統已啟動，輸入 quit,exit,bye 離開")

    while True:
        user_input = input("你：").strip()

        if not user_input:
            continue

        if user_input.lower() in ["quit", "exit", "bye"]:
            print("👋 再見，祝學習愉快！")
            break

        response = chat_with_ollama(user_input)
        print("家教老師：", response)
